#import xlwt,xlrd
import argparse
import os
import pandas as pd
from openpyxl import Workbook
import xlwt
import time
import io
##testpycharm
#################################################################sub function
workbook = Workbook()
workbook.active 
# = wb.activeorkbook = xlwt.Workbook(encoding = 'utf-8')
#style = xlwt.easyxf('font:name Times New Roman')
work_sheets = {}

def parseFile(infile):
    outlist = []
    for line in io.open(infile,'r',encoding='utf-8'):
        tmp = line.rstrip().split("\t")
        if not tmp:
            outlist.append(["\n"])
        else:
            outlist.append(tmp)
    return outlist

def writeSheet(worksheet,inlist):
    for index1,e1 in enumerate(inlist):
        worksheet.append(e1)
	#print (e1)
	#for index2,e2 in enumerate(e1):
        #    worksheet.write(index1, index2, label = e2, style=style)

def Merge (args):
    if not os.path.exists(args.indir):
        #print ("Please import input directory\n");
        sys.exit()
    if not os.path.exists(args.outdir):
        os.mkdir(args.outdir)

    out = os.path.join(args.outdir, args.prefix+'.xlsx')
    root=args.indir
    files=os.listdir(root)
    
    if not args.list :
        sort_files=files
    else:
        sort_files=args.list.split(",")

    for file in sort_files:
        if file not in files:
            exit("list %s not in %s" % ( file, root))
        pathfile = os.path.join(root,file)
        filename = os.path.splitext(file)[0]
        if not os.path.getsize(pathfile):
            file_content=pd.DataFrame(columns = [''])
        elif file.endswith('.txt'):
            olist = parseFile(pathfile)
            #wksheet = workbook.add_sheet(filename)
            wksheet = workbook.create_sheet(filename)
            writeSheet(wksheet,olist)
        # elif file.endswith('.xls'):
        #     file_content = pd.read_excel(pathfile,sep = "\t",header = None)
        # elif file.endswith('.xlsx'):
        #     file_content = pd.read_excel(pathfile,sep = "\t",header = None)
        else:
            continue
    workbook.save(out)


def Split (args):
    if not os.path.exists(args.outdir):
        os.mkdir(args.outdir)
    sheet_list = pd.read_excel(args.excel,sheet_name = None, index_col = 0 ,keep_default_na=False,
                               header = None)
    sep="\t"
    suffix=".txt"
    if args.type == 'csv':
        sep=","
        suffix=".csv"
    for num in sheet_list:
        time.sleep(0.1)
        sheet_list[num].to_csv(os.path.join(args.outdir , num+suffix), 
                               encoding = 'utf-8',sep = sep,header = False )


################################################################# parameter
Parser = argparse.ArgumentParser(description="File split and merge\n")
subparsers = Parser.add_subparsers(help = "Create sub command")

Parser_a = subparsers.add_parser('Merge',help = "Merge file")
Parser_a.add_argument("-I","--indir",help="input directory\n")
Parser_a.add_argument("-L","--list",help="sorted output sheet list\n")
Parser_a.add_argument("-O","--outdir",help="output directory\n")
Parser_a.add_argument("-P","--prefix",help="output file prefix\n")
Parser_a.set_defaults(func = Merge)

Parser_s = subparsers.add_parser('Split',help = "Split file")
Parser_s.add_argument("-e","--excel",help="import excel file\n")
Parser_s.add_argument("-o","--outdir",help="output directory\n")
Parser_s.add_argument("-t","--type",help="output format,csv|table\n")
Parser_s.set_defaults(func = Split)

args = Parser.parse_args()
args.func(args)

